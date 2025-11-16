from fastapi import APIRouter, HTTPException
from typing import List, Dict, Any
from pydantic import BaseModel
from pathlib import Path
import shutil
import yaml
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .. import main as state
from ..core.models import OwnerRecord
from ..core.utils import dump_yaml_entities

router = APIRouter(prefix="/api", tags=["owners"])


@router.get("/owners", response_model=List[OwnerRecord])
async def list_owners():
    return list(state.OWNER_DB.values())


@router.post("/owners", response_model=OwnerRecord, status_code=201)
async def add_owner(payload: OwnerRecord):
    key = payload.name.strip().lower()
    if not state.ALNUM_UNDERSCORE_LOWER_RE.match(key):
        raise HTTPException(status_code=400, detail="Invalid owner name: lowercase alphanumeric and underscore only")
    if key in state.OWNER_DB:
        raise HTTPException(status_code=409, detail="Owner already exists")
    state.OWNER_DB[key] = {
        "name": key,
        "bankaccounts": [b.strip().lower() for b in payload.bankaccounts],
        "properties": [p.strip().lower() for p in payload.properties],
        "companies": [c.strip().lower() for c in payload.companies],
        "export_dir": (payload.export_dir or "").strip(),
    }
    # persist YAML
    try:
        if state.OWNERS_CSV_PATH:
            dump_yaml_entities(state.OWNERS_CSV_PATH.with_suffix('.yaml'), list(state.OWNER_DB.values()), key_field='name')
    except Exception:
        pass
    return state.OWNER_DB[key]


@router.delete("/owners/{name}", status_code=204)
async def delete_owner(name: str):
    key = name.strip().lower()
    if key not in state.OWNER_DB:
        raise HTTPException(status_code=404, detail="Owner not found")
    del state.OWNER_DB[key]
    # persist YAML
    try:
        if state.OWNERS_CSV_PATH:
            dump_yaml_entities(state.OWNERS_CSV_PATH.with_suffix('.yaml'), list(state.OWNER_DB.values()), key_field='name')
    except Exception:
        pass
    return


class OwnerExportPayload(BaseModel):
    name: str


def _read_yaml_map(path: Path) -> Dict[str, Any]:
    try:
        with path.open('r', encoding='utf-8') as f:
            data = yaml.safe_load(f) or {}
            return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _read_processed_rows(processed_dir: Path, ba: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    csv_path = processed_dir / f"{ba}.csv"
    yaml_path = processed_dir / f"{ba}.yaml"
    if csv_path.exists():
        try:
            with csv_path.open('r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for r in reader:
                    rows.append(dict(r))
        except Exception:
            pass
    elif yaml_path.exists():
        try:
            with yaml_path.open('r', encoding='utf-8') as f:
                data = yaml.safe_load(f) or []
                if isinstance(data, list):
                    for it in data:
                        if isinstance(it, dict):
                            rows.append(it)
        except Exception:
            pass
    return rows


def _export_one_owner(owner_name: str) -> Dict[str, Any]:
    if not state.ACCOUNTS_DIR_PATH or not state.CURRENT_YEAR:
        return {"owner": owner_name, "status": "error", "error": "ACCOUNTS_DIR or CURRENT_YEAR not configured"}

    processed_dir: Path = state.PROCESSED_DIR_PATH
    if not processed_dir:
        return {"owner": owner_name, "status": "error", "error": "Processed directory not configured"}

    rentals_dir: Path = state.ACCOUNTS_DIR_PATH / state.CURRENT_YEAR / 'rentalsummary'
    company_dir: Path = state.ACCOUNTS_DIR_PATH / state.CURRENT_YEAR / 'companysummary'

    owner = (state.OWNER_DB or {}).get(owner_name)
    if not owner:
        return {"owner": owner_name, "status": "error", "error": "Owner not found"}
    export_dir = (owner.get('export_dir') or '').strip()
    if not export_dir:
        return {"owner": owner_name, "status": "error", "error": "export_dir not set"}

    base_dir: Path = Path(export_dir).expanduser().resolve()
    if not base_dir.exists() or not base_dir.is_dir():
        return {"owner": owner_name, "status": "error", "error": "export_dir does not exist or is not a directory"}

    dest_root: Path = base_dir / state.CURRENT_YEAR / 'export'
    try:
        dest_root.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        return {"owner": owner_name, "status": "error", "error": f"mkdir failed: {e}"}

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    # Rental summary sheet (one row per property, fixed ordered columns)
    ws_rent = wb.create_sheet(title="RentalSummary")
    rent_cols = [
        "property","rent","commissions","insurance","proffees","mortgageinterest",
        "repairs","tax","utilities","depreciation","hoa","other","costbasis","renteddays","profit"
    ]
    ws_rent.append(rent_cols)
    for prop in (owner.get('properties') or []):
        p = (prop or '').strip().lower()
        if not p:
            continue
        p_yaml = rentals_dir / f"{p}.yaml"
        row_vals = {k: '' for k in rent_cols}
        row_vals['property'] = p
        if p_yaml.exists():
            data = _read_yaml_map(p_yaml)
            for k in rent_cols:
                if k == 'property':
                    continue
                try:
                    v = data.get(k)
                    if v is None:
                        continue
                    row_vals[k] = float(v)
                except Exception:
                    row_vals[k] = data.get(k)
        ws_rent.append([row_vals.get(c, '') for c in rent_cols])
    # Autofilter and best-fit widths
    if ws_rent.max_row >= 1 and ws_rent.max_column >= 1:
        ws_rent.auto_filter.ref = f"A1:{get_column_letter(ws_rent.max_column)}{ws_rent.max_row}"
    for col_idx in range(1, ws_rent.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws_rent.max_row + 1):
            val = ws_rent.cell(row=row, column=col_idx).value
            s = str(val) if val is not None else ''
            if len(s) > max_len:
                max_len = len(s)
        ws_rent.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    # Company summary sheet (one row per company, fixed ordered columns)
    ws_comp = wb.create_sheet(title="CompanySummary")
    comp_cols = [
        "Name","income","rentpassedtoowners","bankfees","c_auto","c_donate","c_entertainment",
        "c_internet","c_license","c_mobile","c_off_exp","c_parktoll","c_phone","c_website",
        "ignore","insurane","proffees","utilities","profit"
    ]
    ws_comp.append(comp_cols)
    for comp in (owner.get('companies') or []):
        c = (comp or '').strip().lower()
        if not c:
            continue
        row_vals = {k: '' for k in comp_cols}
        row_vals['Name'] = c
        c_yaml = company_dir / f"{c}.yaml"
        if c_yaml.exists():
            data = _read_yaml_map(c_yaml)
            for k in comp_cols:
                if k == 'Name':
                    continue
                try:
                    v = data.get(k)
                    if v is None:
                        continue
                    row_vals[k] = float(v)
                except Exception:
                    row_vals[k] = data.get(k)
        ws_comp.append([row_vals.get(cn, '') for cn in comp_cols])
    # Autofilter and best-fit widths
    if ws_comp.max_row >= 1 and ws_comp.max_column >= 1:
        ws_comp.auto_filter.ref = f"A1:{get_column_letter(ws_comp.max_column)}{ws_comp.max_row}"
    for col_idx in range(1, ws_comp.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws_comp.max_row + 1):
            val = ws_comp.cell(row=row, column=col_idx).value
            s = str(val) if val is not None else ''
            if len(s) > max_len:
                max_len = len(s)
        ws_comp.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    # Per-bankaccount sheets
    for ba in (owner.get('bankaccounts') or []):
        ban = (ba or '').strip().lower()
        if not ban:
            continue
        rows = _read_processed_rows(processed_dir, ban)
        if not rows:
            continue
        ws_ba = wb.create_sheet(title=(ban[:31]))
        header = list(rows[0].keys())
        ws_ba.append(header)
        for r in rows:
            ws_ba.append([r.get(h, '') for h in header])
        # Autofilter and best-fit widths
        if ws_ba.max_row >= 1 and ws_ba.max_column >= 1:
            ws_ba.auto_filter.ref = f"A1:{get_column_letter(ws_ba.max_column)}{ws_ba.max_row}"
        for col_idx in range(1, ws_ba.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in range(1, ws_ba.max_row + 1):
                val = ws_ba.cell(row=row, column=col_idx).value
                s = str(val) if val is not None else ''
                if len(s) > max_len:
                    max_len = len(s)
            ws_ba.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    out_path = dest_root / f"export_{state.CURRENT_YEAR}.xlsx"
    try:
        wb.save(str(out_path))
        return {"owner": owner_name, "status": "ok", "path": str(out_path)}
    except Exception as e:
        return {"owner": owner_name, "status": "error", "error": f"save failed: {e}"}


@router.post("/owners/export")
async def export_owner(payload: OwnerExportPayload):
    name = (payload.name or '').strip().lower()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    result = _export_one_owner(name)
    if result.get("status") != "ok":
        # Still return 200 with error details to keep UI simple
        return result
    return result


@router.post("/owners/export-all")
async def export_all_owners():
    results: List[Dict[str, Any]] = []
    for name in (state.OWNER_DB or {}).keys():
        res = _export_one_owner(name)
        if res.get("status") == "ok" or res.get("error") != "Owner not found":
            results.append(res)
    return {"status": "ok", "results": results}


class OwnerPrepPayload(BaseModel):
    name: str = ""


@router.post("/owners/prepentities")
async def prep_entities(payload: OwnerPrepPayload):
    name = (payload.name or "").strip().lower()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    owner = (state.OWNER_DB or {}).get(name)
    if not owner:
        raise HTTPException(status_code=404, detail="Owner not found")

    if not state.ACCOUNTS_DIR_PATH or not state.CURRENT_YEAR:
        raise HTTPException(status_code=500, detail="ACCOUNTS_DIR or CURRENT_YEAR not configured")

    entities_dir: Path = f'{state.ACCOUNTS_DIR_PATH}/{state.CURRENT_YEAR}/entities' 
    entities_dir = Path(entities_dir).expanduser().resolve()

    if not entities_dir or not entities_dir.exists():
        raise HTTPException(status_code=500, detail="entities dir is not resolved")

    dest_root: Path = state.ACCOUNTS_DIR_PATH / state.CURRENT_YEAR / 'export' / name / 'entities'

    # Copy top-level YAMLs to entities/
    try:
        for fname in ['banks.yaml', 'transaction_types.yaml','tax_category.yaml','inherit_common_to_bank.yaml', 'common_rules.yaml']:
            src = entities_dir / fname
            if src.exists() and src.is_file():
                shutil.copy2(src, dest_root / fname)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed copying top-level YAMLs: {e}")

    # Helpers
    def _load_yaml_list(path: Path) -> List[Dict[str, Any]]:
        try:
            with path.open('r', encoding='utf-8') as f:
                data = yaml.safe_load(f) or []
                return data if isinstance(data, list) else []
        except Exception:
            return []

    def _save_yaml_list(path: Path, rows: List[Dict[str, Any]]):
        try:
            with path.open('w', encoding='utf-8') as f:
                yaml.safe_dump(rows, f, sort_keys=True, allow_unicode=True)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed writing {path.name}: {e}")

    # companies.yaml
    try:
        comp_src = entities_dir / 'companies.yaml'
        all_comp = _load_yaml_list(comp_src)
        want = set([ (c or '').strip().lower() for c in (owner.get('companies') or []) ])
        filtered: List[Dict[str, Any]] = []
        for rec in all_comp:
            if not isinstance(rec, dict):
                continue
            key = (str(rec.get('companyname') or '')).strip().lower()
            if key and key in want:
                filtered.append(rec)
        _save_yaml_list(dest_root / 'companies.yaml', filtered)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed filtering companies.yaml: {e}")

    # properties.yaml
    try:
        prop_src = entities_dir / 'properties.yaml'
        all_props = _load_yaml_list(prop_src)
        want = set([ (p or '').strip().lower() for p in (owner.get('properties') or []) ])
        filtered = []
        for rec in all_props:
            if not isinstance(rec, dict):
                continue
            key = (str(rec.get('property') or '')).strip().lower()
            if key and key in want:
                filtered.append(rec)
        _save_yaml_list(dest_root / 'properties.yaml', filtered)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed filtering properties.yaml: {e}")

    # bankaccounts.yaml
    try:
        ba_src = entities_dir / 'bankaccounts.yaml'
        all_bas = _load_yaml_list(ba_src)
        want = set([ (b or '').strip().lower() for b in (owner.get('bankaccounts') or []) ])
        filtered = []
        for rec in all_bas:
            if not isinstance(rec, dict):
                continue
            key = (str(rec.get('bankaccountname') or '')).strip().lower()
            if key and key in want:
                filtered.append(rec)
        _save_yaml_list(dest_root / 'bankaccounts.yaml', filtered)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed filtering bankaccounts.yaml: {e}")

    # owners.yaml (filtered to just this owner)
    try:
        owners_src = entities_dir / 'owners.yaml'
        all_owners = _load_yaml_list(owners_src)
        filtered = []
        for rec in all_owners:
            if not isinstance(rec, dict):
                continue
            key = (str(rec.get('name') or '')).strip().lower()
            if key == name:
                filtered.append(rec)
        _save_yaml_list(dest_root / 'owners.yaml', filtered)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed filtering owners.yaml: {e}")

    return {"status": "ok", "owner": name, "export_path": str(dest_root)}


@router.post("/export-accounts")
async def export_accounts():
    """Export accounts for all owners that have export_dir set.
    Mirrors behavior requested for /api/export-accounts.
    """
    results: List[Dict[str, Any]] = []
    for name, owner in (state.OWNER_DB or {}).items():
        if not (owner.get('export_dir') or '').strip():
            continue
        res = _export_one_owner(name)
        results.append(res)
    return {"status": "ok", "results": results}
